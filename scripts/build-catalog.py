#!/usr/bin/env python3
"""Read the finalized spreadsheet, reconcile it with the image folders,
and emit src/data/catalog.json — the single source of truth for the site.

The spreadsheet is authoritative for category/product/URL. The image folders
are authoritative for which files exist. Any disagreement is reported, never
silently patched.
"""
import os, json, re, sys
from openpyxl import load_workbook

ASSETS = "/Users/sajjadahmad/Documents/the-retail-packaging"
XLSX = os.path.join(ASSETS, "categories-products-final.xlsx")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "src", "data", "catalog.json")
REPORT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                      "reports", "image-validation.json")
SYSTEM = {".DS_Store", ".localized", "Thumbs.db"}
IMG_EXT = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".avif"}


def kids(p):
    return sorted((n for n in os.listdir(p) if not n.startswith("._") and n not in SYSTEM),
                  key=str.lower)


def slugify(s):
    return re.sub(r"[^a-z0-9]+", "-", s.strip().lower()).strip("-")


wb = load_workbook(XLSX, data_only=True)

# ---- Categories sheet -------------------------------------------------
cat_rows = []
ws = wb["Categories"]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1] or r[1] == "TOTAL":
        continue
    cat_rows.append({"name": str(r[1]).strip(), "url": str(r[2]).strip(),
                     "products": r[3], "images": r[4]})

# ---- Products sheet (authoritative) -----------------------------------
prod_rows = []
ws = wb["Products"]
head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[1]:
        continue
    prod_rows.append({
        "category": str(r[1]).strip(),
        "categoryUrl": str(r[2]).strip(),
        "product": str(r[3]).strip(),
        "productUrl": str(r[4]).strip(),
        "images": int(r[5] or 0),
    })

# ---- Removed sheet (never publish) ------------------------------------
removed = set()
if "Removed" in wb.sheetnames:
    for r in wb["Removed"].iter_rows(min_row=2, values_only=True):
        if r and r[1]:
            removed.add(slugify(str(r[1])))

# ---- Disk -------------------------------------------------------------
disk = {}
for c in kids(ASSETS):
    cp = os.path.join(ASSETS, c)
    if not os.path.isdir(cp) or c.startswith("_") or c.startswith("."):
        continue
    for p in kids(cp):
        pp = os.path.join(cp, p)
        if not os.path.isdir(pp):
            continue
        files = [f for f in kids(pp) if os.path.isfile(os.path.join(pp, f))]
        disk[(c, slugify(p))] = {
            "dir": pp, "folder": p, "category": c,
            "images": [f for f in files if os.path.splitext(f)[1].lower() in IMG_EXT],
            "other": [f for f in files if os.path.splitext(f)[1].lower() not in IMG_EXT],
        }

# ---- Reconcile --------------------------------------------------------
issues = {"missingOnDisk": [], "extraOnDisk": [], "countMismatch": [],
          "noImages": [], "removedLeak": [], "duplicateFiles": []}

catalog_cats = {}
for c in cat_rows:
    catalog_cats[c["url"].strip("/")] = {
        "name": c["name"], "slug": c["url"].strip("/"),
        "url": "/" + c["url"].strip("/") + "/", "products": [],
    }

products = []
seen_urls = {}
for p in prod_rows:
    cslug = p["categoryUrl"].strip("/")
    pslug = p["productUrl"].strip("/")
    key = (cslug, pslug)
    if pslug in removed:
        issues["removedLeak"].append(pslug)
        continue
    if pslug in seen_urls:
        issues["duplicateFiles"].append(f"duplicate product URL /{pslug}/")
    seen_urls[pslug] = cslug

    d = disk.get(key)
    if not d:
        issues["missingOnDisk"].append(f"{cslug}/{p['product']}")
        imgs = []
        folder = None
    else:
        imgs = d["images"]
        folder = d["folder"]
        if len(imgs) != p["images"]:
            issues["countMismatch"].append(
                {"product": pslug, "sheet": p["images"], "disk": len(imgs)})
    if not imgs:
        issues["noImages"].append(pslug)

    rec = {
        "name": p["product"], "slug": pslug, "url": f"/{pslug}/",
        "category": p["category"], "categorySlug": cslug,
        "categoryUrl": f"/{cslug}/",
        "sourceDir": d["dir"] if d else None,
        "sourceFolder": folder,
        "imageFiles": imgs,
        "imageCount": len(imgs),
    }
    products.append(rec)
    if cslug in catalog_cats:
        catalog_cats[cslug]["products"].append(pslug)

sheet_keys = {(p["categoryUrl"].strip("/"), p["productUrl"].strip("/")) for p in prod_rows}
for k, v in disk.items():
    if k not in sheet_keys:
        issues["extraOnDisk"].append(f"{k[0]}/{v['folder']}")

catalog = {
    "brand": {
        "name": "The Retail Packaging",
        "domain": "https://theretailpackaging.com",
        "whatsapp": "15033580443",
        "whatsappDisplay": "+1 503-358-0443",
    },
    "categories": list(catalog_cats.values()),
    "products": products,
    "totals": {
        "categories": len(catalog_cats),
        "products": len(products),
        "images": sum(p["imageCount"] for p in products),
    },
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
os.makedirs(os.path.dirname(REPORT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(catalog, f, indent=1, ensure_ascii=False)
with open(REPORT, "w", encoding="utf-8") as f:
    json.dump({"totals": catalog["totals"], "issues": issues}, f, indent=1, ensure_ascii=False)

print(f"categories : {catalog['totals']['categories']}")
print(f"products   : {catalog['totals']['products']}")
print(f"images     : {catalog['totals']['images']}")
print(f"removed listed: {len(removed)}")
for k, v in issues.items():
    print(f"  {k:<16} {len(v)}" + (f"  -> {v[:4]}" if v else ""))
print(OUT)
